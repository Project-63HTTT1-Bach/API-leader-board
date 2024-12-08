import sqlite3
from openpyxl import load_workbook
import os
import bcrypt
from datetime import datetime, timedelta

excel_file_1 = "Danh sách điểm danh HTKDTM 63HTTT1.xlsx"
excel_file_2 = "Danh_sach_tich_cuc.xlsx"

for file in [excel_file_1, excel_file_2]:
    if not os.path.exists(file):
        print(f"File '{file}' không tồn tại.")
        exit()
    else:
        print(f"File '{file}' tồn tại.")

password = "admin"  
hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

conn = sqlite3.connect("student_management.db")
cursor = conn.cursor()

workbook_1 = load_workbook(filename=excel_file_1, data_only=True)
sheet_1 = workbook_1.active  
rows = list(sheet_1.iter_rows(values_only=True))
data = rows[9:68] 
# print(f"Data từ {excel_file_1}: {data}")

workbook_2 = load_workbook(filename=excel_file_2, data_only=True)
# sheets = workbook_2.sheetnames
# print(f"File Excel có các sheet: {sheets}")

sheet2_name = "DS_DiemDanh_15cot" 
sheet2 = workbook_2[sheet2_name]
rows1 = list(sheet2.iter_rows(values_only=True))
data1 = rows1[2:22]
# print(f"Data1 từ {sheet2_name}: {data1}")

sheet3_name = "Điểm cộng cụm" 
sheet3 = workbook_2[sheet3_name]
rows2 = list(sheet3.iter_rows(values_only=True))
data2 = rows2[1:4]  
# print(f"Data2 từ {sheet3_name}: {data2}")

# Thêm users vào database
cursor.execute("""
INSERT OR REPLACE INTO Users (user_id, role, password)
VALUES (?, ?, ?)
""", ('admin', 0, hashed_password))

for row in data:
    cursor.execute("""
    INSERT INTO Users (user_id, role)
    VALUES (?, ?)
    """, (row[1], 1))


# Thêm students vào database
for row in data1:
    cursor.execute("""
    INSERT INTO Students (student_id, full_name, cluster_number, group_number)
    VALUES (?, ?, ?, ?)
    """, (row[1], row[2] + " " + row[3], 1, row[5]))

for row in data:
    cursor.execute("""SELECT COUNT(*) FROM Students WHERE student_id = ?""", (row[1],))
    count = cursor.fetchone()[0]

    if count == 0:  
        cursor.execute("""
        INSERT INTO Students (student_id, full_name)
        VALUES (?, ?)
        """, (row[1], row[2] + " " + row[3]))

# Thêm điểm danh vào database
for row in data:
    student_id = row[1]
    attendance_data = row[5:21]
    start_date = datetime(2024, 11, 12)  # Ngày bắt đầu từ 12/11/2024 (thứ 3)
    end_date = datetime(2024, 12, 6)  # Giới hạn ngày cuối là 06/12/2024
    
    # matched_row = next((r for r in data1 if r[1] == student_id), None)
    
    # if matched_row:
    for i, status in enumerate(attendance_data):
        current_date = start_date + timedelta(weeks=i//2, days=(i % 2) * 3)
        
        if current_date > end_date:
            break
        
        status_code = 1 if status == 'pb' or status is None else 0 if status == 'v' else None
        
        if status_code is not None:  
            cursor.execute("""
            INSERT INTO Attendance (student_id, date, status)
            VALUES (?, ?, ?)
            """, (student_id, current_date.date(), status_code))


# Thêm điểm tích cực vào database
current_date = datetime.now().date()
reasons = {
    6: "Lên bảng",
    7: "Mindmap tổng hợp",
    8: "Code hệ thống",
    10: "Không mang lap"  # Đây là điểm trừ
}
for row in data1:
    student_id = row[1]

    for col_index, reason in reasons.items():
        activity = row[col_index]

        if activity:  
            points = activity.count('x')
            
            if reason == "Không mang lap":
                points = -points  
            
            if points != 0:
                cursor.execute("""
                INSERT INTO BonusPoints (student_id, reason, points, awarded_date)
                VALUES (?, ?, ?, ?)
                """, (student_id, reason, points, current_date))

for event in data2:
    event_date = event[0].date()  # Lấy ngày sự kiện
    reason = event[1]  # Lý do sự kiện

    for row in data1:
        student_id = row[1] 

        cursor.execute("""
        INSERT INTO BonusPoints (student_id, reason, points, awarded_date)
        VALUES (?, ?, ?, ?)
        """, (student_id, reason, 1, event_date))
    
conn.commit()
conn.close()